import tkinter
import pygame
import random
import time
import openpyxl
from openpyxl import Workbook
#颜色
white=(255,255,255)
black=(0,0,0)
yellow=(255,255,0)
Blue=(0,0,255)
DarkBlue=(0,0,128)
skin=(248,248,255)


#常数
pos_snake=[[250,250],[240,250],[230,250],[220,250]]#蛇的位置
pos_snake_head=[250,250]
pos_food=[random.randrange(1,50)*10,random.randrange(1,50)*10]#随机生成果实的位置
screen_width=500#屏幕宽度
screen_height=500#屏幕长度
cell=10#蛇的单位长度
speed=5#自动移动的速度，即每秒自动移动的次数，可以取小数





def main():#构建初始界面
    #主窗口
    win=tkinter.Tk()
    win.title("贪吃蛇1.3")
    win.geometry("500x600")

    #组件

    #标题
    mainname=tkinter.Label(win,text="贪 吃 蛇",font=("Arial",30))
    mainname.place(relx=0,relwidth=1,rely=0,relheight=0.25)


    #按钮
    b1=tkinter.Button(win,text="开始游戏",bg='Blue',fg='white',activebackground='DarkBlue',activeforeground='white',command=play)
    b1.place(relx=0.25,relwidth=0.5,rely=0.25,relheight=0.1)
    b3=tkinter.Button(win,text="设置",bg='Blue',fg='white',activebackground='DarkBlue',activeforeground='white',command=setting)
    b3.place(relx=0.25,relwidth=0.5,rely=0.4,relheight=0.1)
    b4=tkinter.Button(win,text="排行榜",bg='Blue',fg='white',activebackground='DarkBlue',activeforeground='white',command=readrank)
    b4.place(relx=0.25,relwidth=0.5,rely=0.55,relheight=0.1)
    b5=tkinter.Button(win,text="退出游戏",bg='Blue',fg='white',activebackground='DarkBlue',activeforeground='white',command=win.destroy)
    b5.place(relx=0.25,relwidth=0.5,rely=0.7,relheight=0.1)
    win.mainloop()

def play():#用pygame进行游戏（单人游戏）
    global pos_snake,pos_snake_head,speed,skin
    pygame.init()#初始化
    pygame.display.set_caption("贪吃蛇1.0")#设置标题
    s=pygame.display.set_mode((screen_height,screen_width))#设置游戏窗口对象
    score_font=pygame.font.SysFont('方正粗黑宋简体',15)#创建字体对象，用来打印分数
    title_font=pygame.font.SysFont('方正粗黑宋简体',50)#标题所用字体
    menu_font=pygame.font.SysFont('方正粗黑宋简体',20)#菜单栏所用字体
       
    
    #下面是游戏过程
    direct=3
    clock=pygame.time.Clock()
    for pos in pos_snake:#画出蛇的初始位置
        draw(s,skin,pos)
    draw(s,yellow,pos_food)#画出初始的果实
    while True:
        for event in pygame.event.get():#检测键盘操作
            if event.type==pygame.KEYDOWN:#如果按下了键
                if event.key==pygame.K_UP:
                    direct=0
                elif event.key==pygame.K_DOWN:
                    direct=1
                elif event.key==pygame.K_LEFT:
                    direct=2
                elif event.key==pygame.K_RIGHT:
                    direct=3
        move(direct)
        s.fill(black)
        score=round((len(pos_snake)-4)*speed/8,2)
        printtext(s,"当前分数为："+str(score),60,20,score_font)
        printtext(s,"当前速度为："+str(speed),60,40,score_font)
        printtext(s,"分数为基础得分乘速度除8",85,60,score_font)
        for pos in pos_snake:#画出蛇的位置
            draw(s,skin,pos)
        draw(s,yellow,pos_food)
        if hit1() or hit2():
            printtext(s,"游戏结束",250,250,title_font)
            printtext(s,"您的最终得分为"+str(score),245,300,menu_font)
            printtext(s,"5秒钟后自动返回菜单",245,350,menu_font)  
            pygame.display.update()
            pos_snake=[[250,250],[240,250],[230,250],[220,250]]#蛇的位置
            pos_snake_head=[250,250]
            break
        clock.tick(speed)#设置游戏帧率
        pygame.display.update()
    time.sleep(5)
    pygame.quit()
    
    #下面是根据成绩来保存用户信息
    fn = r'rank.xlsx'
    wb = openpyxl.load_workbook(fn)
    ws = wb.worksheets[0]

    if (score > ws['B5'].value):
        ws['B5'].value = score
        # 此处需要加一个GUI弹窗输入姓名赋值给ws['A5'].value
        win = tkinter.Toplevel()
        win.title("成绩采集")
        win.geometry("300x200")

        def s_save():
            if e.get()=="":
                user="无名氏"
            else:
                user = e.get()
            for a in range(ws.max_row, 0, -1):  # we.rows 获取每一行数据
                for i in range(ws.max_column, 0, -1):
                    if (i == ws.max_column):
                        if (score > ws.cell(a, i).value):
                            if (a != ws.max_row):
                                ws.cell(a + 1, i - 1).value = ws.cell(a, i - 1).value
                                ws.cell(a + 1, i).value = ws.cell(a, i).value
                                ws.cell(a, i - 1).value = user
                                ws.cell(a, i).value = score
                            else:
                                ws.cell(a, i - 1).value = user
                                ws.cell(a, i).value = score
                        wb.save(fn)
            win.destroy()

        e = tkinter.StringVar()  # 用来关联文本的量
        w = tkinter.Entry(win, textvariable=e).place(relx=0.1, relwidth=0.8, rely=0.25, relheight=0.2)

        b2 = tkinter.Button(win, text="退出", bg='Blue', fg='white', activebackground='DarkBlue',
                            activeforeground='white', command=s_save)
        b2.place(relx=0.2, relwidth=0.6, rely=0.75, relheight=0.2)

        win.mainloop()

    

def setting():#设置窗口
    s=tkinter.Tk()
    s.title("设置")
    s.geometry("500x600")

    #组件

    #标题
    mainname=tkinter.Label(s,text="设 置",font=("Arial",30))
    mainname.place(relx=0,relwidth=1,rely=0,relheight=0.25)


    #按钮
    b1=tkinter.Button(s,text="修改速度",bg='Blue',fg='white',activebackground='DarkBlue',activeforeground='white',command=set_speed)
    b1.place(relx=0.25,relwidth=0.5,rely=0.4,relheight=0.1)
    b3=tkinter.Button(s,text="选择皮肤",bg='Blue',fg='white',activebackground='DarkBlue',activeforeground='white',command=set_skin)
    b3.place(relx=0.25,relwidth=0.5,rely=0.55,relheight=0.1)
    b4=tkinter.Button(s,text="返回主界面",bg='Blue',fg='white',activebackground='DarkBlue',activeforeground='white',command=s.destroy)
    b4.place(relx=0.25,relwidth=0.5,rely=0.7,relheight=0.1)
    s.mainloop()
    
def readrank():#排行榜窗口
    fn = r'rank.xlsx'
    wb=openpyxl.load_workbook(fn)
    ws=wb.worksheets[0]
    
    sk=tkinter.Tk()
    sk.title("排行榜")
    sk.geometry("500x600")
    
    mainname=tkinter.Label(sk,text="排 行 榜",font=("Arial",30))
    mainname.place(relx=0,relwidth=1,rely=0,relheight=0.25)
    
    s1=tkinter.Label(sk,font=("Arial",15),text="1."+str(ws['A1'].value)+"  分数为"+str(ws['B1'].value))
    s1.place(relx=0.2,relwidth=0.6,rely=0.15,relheight=0.1)
    s2=tkinter.Label(sk,font=("Arial",15),text="2."+str(ws['A2'].value)+"  分数为"+str(ws['B2'].value))
    s2.place(relx=0.2,relwidth=0.6,rely=0.3,relheight=0.1)
    s3=tkinter.Label(sk,font=("Arial",15),text="3."+str(ws['A3'].value)+"  分数为"+str(ws['B3'].value))
    s3.place(relx=0.2,relwidth=0.6,rely=0.45,relheight=0.1)
    s4=tkinter.Label(sk,font=("Arial",15),text="4."+str(ws['A4'].value)+"  分数为"+str(ws['B4'].value))
    s4.place(relx=0.2,relwidth=0.6,rely=0.6,relheight=0.1)
    s5=tkinter.Label(sk,font=("Arial",15),text="5."+str(ws['A5'].value)+"  分数为"+str(ws['B5'].value))
    s5.place(relx=0.2,relwidth=0.6,rely=0.75,relheight=0.1)
    b4=tkinter.Button(sk,text="返回主界面",bg='Blue',fg='white',activebackground='DarkBlue',activeforeground='white',command=sk.destroy)
    b4.place(relx=0.25,relwidth=0.5,rely=0.88,relheight=0.1)
    sk.mainloop()


def set_speed():
    global speed
    ss=tkinter.Tk()
    ss.title("设置速度")
    ss.geometry("500x600")
    l=tkinter.Label(ss,font=("Arial",30))
    l.place(relx=0,relwidth=1,rely=0,relheight=0.25)
    def print_speed(v):
        global speed
        l.config(text="速度为"+v,font=("Arial",10))
        speed=float(v)
    hua= tkinter.Scale(ss, label='滑动滑条选择速度（即每秒刷新次数）', from_=4, to=12, orient=tkinter.HORIZONTAL, length=200, showvalue=0,tickinterval=2, resolution=0.01, command=print_speed)
    hua.pack()
    b1=tkinter.Button(ss,text="退出",bg='Blue',fg='white',activebackground='DarkBlue',activeforeground='white',command=ss.destroy)
    b1.place(relx=0.25,relwidth=0.5,rely=0.4,relheight=0.1)
    ss.mainloop()


def set_skin():
    global skin
    
    sk=tkinter.Toplevel()
    sk.title("皮肤选择")
    sk.geometry("500x600")
    
    mainname=tkinter.Label(sk,text="设 置",font=("Arial",30))
    mainname.place(relx=0,relwidth=1,rely=0,relheight=0.25)
    
    var=tkinter.StringVar()
    #调试用   
    l = tkinter.Label(sk, bg='white', width=40, text='empty')
    l.place(relx=0.25,relwidth=0.5,rely=0.18,relheight=0.1)
    
    def choose_skin():
        global skin

        if var.get()=="幽灵鬼魅":
            skin=(248,248,255)
        elif var.get()=="春之礼赞":
            skin=(124,252,0)
        elif var.get()=="水天一色":
            skin=(0,191,255)
        elif var.get()=="骄阳似火":
            skin=(255,69,0)
        #调试用
        l.config(text='当前选择为皮肤' + var.get())
    r1=tkinter.Radiobutton(sk,text="幽灵鬼魅",variable=var,value="幽灵鬼魅",command=choose_skin)
    r1.place(relx=0.25,relwidth=0.5,rely=0.3,relheight=0.1)
    r2=tkinter.Radiobutton(sk,text="春之礼赞",variable=var,value="春之礼赞",command=choose_skin)
    r2.place(relx=0.25,relwidth=0.5,rely=0.45,relheight=0.1)
    r3=tkinter.Radiobutton(sk,text="水天一色",variable=var,value="水天一色",command=choose_skin)
    r3.place(relx=0.25,relwidth=0.5,rely=0.6,relheight=0.1)
    r4=tkinter.Radiobutton(sk,text="骄阳似火",variable=var,value="骄阳似火",command=choose_skin)
    r4.place(relx=0.25,relwidth=0.5,rely=0.75,relheight=0.1)
    b1=tkinter.Button(sk,text="退出",bg='Blue',fg='white',activebackground='DarkBlue',activeforeground='white',command=sk.destroy)
    b1.place(relx=0.25,relwidth=0.5,rely=0.9,relheight=0.1)
    sk.mainloop()


#后面几个是pygame的函数
def draw(s,color,pos):#用来画一格的基本操作
    pygame.draw.rect(s,color,pygame.Rect(pos[0],pos[1],cell,cell))
    
    
def move(direct):#用来操作蛇移动的函数
    global pos_food
    if direct==0:
        pos_snake_head[1]-=cell
    elif direct==1:
        pos_snake_head[1]+=cell
    elif direct==2:
        pos_snake_head[0]-=cell
    elif direct==3:
        pos_snake_head[0]+=cell
    pos_snake.insert(0,list(pos_snake_head))
    if pos_snake_head!=pos_food:#如果没吃到果子尾巴才会收一格，不然不收
        a=pos_snake.pop()
    else:
        pos_food=[random.randrange(1,50)*cell,random.randrange(1,50)*cell]
    print(pos_snake)
    
    
    
def hit1():#用来判断是否撞自己
    if pos_snake_head in pos_snake[1:]:
        return True
    else:
        return False
    
    
def hit2():#用来判断是否撞墙
    if pos_snake_head[0]>screen_width or pos_snake_head[0]<0 or pos_snake_head[1]>screen_height or pos_snake_head[1]<0:
        return True
    else:
        return False
        

def printtext(screen,text,pos1,pos2,the_font):#用来打印文字的函数
    the_Surf=the_font.render(text,True,white,black)
    the_rect=the_Surf.get_rect()
    the_rect.center=(pos1,pos2)
    screen.blit(the_Surf,the_rect)


main()
